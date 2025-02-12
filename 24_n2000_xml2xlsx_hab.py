"""
Author : ExEco Environnement
Edition date : 2025/02
Name : 24_n2000_xml2xlsx_hab
Group : Biblio_PatNat
"""

import xml.etree.ElementTree as ET
import pandas as pd
import os
import sys
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl import load_workbook
from datetime import datetime

def obtain_folder_path():
    return sys.argv[1] if len(sys.argv) > 1 else input("Entrez le chemin du dossier contenant les fichiers XML : ")

def truncate_sheet_name(sheet_name):
    """
    Truncate the sheet name to 31 characters, as required by Excel.
    """
    return sheet_name[:31]  # Truncate to 31 characters

def xml_to_dataframe(xml_file):
    """
    Parse the XML file and return a DataFrame with the contents of the <CD_UE>, <LB_HAB>, and <LB_ZN> tags
    inside the <BIOTOP> elements, but only for rows with <FG_TYPO> == 'D' followed by <CD_UE> and <LB_HAB>.
    """
    try:
        # Parse the XML file
        tree = ET.parse(xml_file)
        root = tree.getroot()

        # Initialize lists to store the data
        cd_habs = []
        lb_habdh_frs = []
        site_name = ""
        sitecode = ""

        # Iterate through <BIOTOP> elements
        for n2000_elem in root.iter('BIOTOP'):
            # Extract SITECODE and SITE_NAME from the current <BIOTOP> element
            sitecode_elem = n2000_elem.find('SITECODE')
            site_name_elem = n2000_elem.find('SITE_NAME')

            if sitecode_elem is not None:
                sitecode = sitecode_elem.text  # Get SITECODE value

            if site_name_elem is not None:
                site_name = site_name_elem.text  # Get SITE_NAME value

            # Iterate through the <TYPO_INFO_ROW> elements
            for typo_info_row_elem in n2000_elem.iter('HABIT1_ROW'):
                cd_hab_elem = typo_info_row_elem.find('CD_UE')
                lb_habdh_fr_elem = typo_info_row_elem.find('LB_HABDH_FR')

                if cd_hab_elem is not None and lb_habdh_fr_elem is not None:
                    cd_habs.append(cd_hab_elem.text)
                    lb_habdh_frs.append(lb_habdh_fr_elem.text)

        # Create a DataFrame from the extracted data
        data = {
            'CD_UE': cd_habs,
            'LB_HABDH_FR': lb_habdh_frs
        }

        df = pd.DataFrame(data)
        return df, site_name, sitecode  # Return SITE_NAME and SITECODE along with the DataFrame

    except ET.ParseError as e:
        print(f"Error parsing XML file {xml_file}: {e}")
        return pd.DataFrame(columns=['CD_UE', 'LB_HABDH_FR']), "", ""
    except Exception as e:
        print(f"Unexpected error with file {xml_file}: {e}")
        return pd.DataFrame(columns=['CD_UE', 'LB_HABDH_FR']), "", ""

def process_xml_files_in_folder(folder_path):
    """
    Process all XML files in a folder and export the data to an Excel file with a summary sheet for each CD_UE.
    """
    # Validate the folder path
    if not os.path.isdir(folder_path):
        print(f"The path {folder_path} is not a valid directory.")
        return

    # List all XML files in the folder
    xml_files = [f for f in os.listdir(folder_path) if f.endswith('.xml') and f.startswith('FR') and len(f) == 13]

    # Initialize sets to store unique CD_UE and LB_HABDH_FR
    unique_cd_ues = set()
    unique_lb_habdh_frs = set()

    # Dictionary to store the presence of LB_HABDH_FR in each file
    hab_presence = {}

    # Get the current date and time for the filename
    current_time = datetime.now().strftime("%Y%m%d%H%M%S")

    # Define the output Excel file path with the timestamp
    excel_file = os.path.join(folder_path, f'output_n2000_xml2xlsx_hab_{current_time}.xlsx')

    # Create an Excel writer object
    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Process each XML file
            for xml_file in xml_files:
                # Construct the full path to the XML file
                full_path = os.path.join(folder_path, xml_file)

                # Convert the XML file to a DataFrame and get SITECODE
                df, site_name, sitecode = xml_to_dataframe(full_path)

                if not df.empty:
                    # Update unique CD_UE and LB_HABDH_FR sets
                    unique_cd_ues.update(df['CD_UE'].unique())
                    unique_lb_habdh_frs.update(df['LB_HABDH_FR'].unique())

                    # Concaténation des valeurs SITECODE et SITE_NAME pour le nom de l'onglet
                    sheet_name = f"{sitecode} - {site_name}"
                    sheet_name_truncated = truncate_sheet_name(sheet_name)  # Tronquer le nom de l'onglet à 31 caractères
                    hab_presence[sheet_name] = set(df['LB_HABDH_FR'])

                    # Write the DataFrame to a new sheet in the Excel file
                    df.to_excel(writer, sheet_name=sheet_name_truncated, index=False)

            # Create the summary table
            summary_data = {
                'CD_UE': list(unique_cd_ues),
                'LB_HABDH_FR': list(unique_lb_habdh_frs)
            }

            # Add columns to the summary for each XML file processed
            for sheet_name in hab_presence:
                summary_data[sheet_name] = ['X' if hab in hab_presence[sheet_name] else '' for hab in unique_lb_habdh_frs]

            # Convert the summary data to a DataFrame
            summary_df = pd.DataFrame(summary_data)

            # Write the summary DataFrame to the Excel file
            summary_df.to_excel(writer, sheet_name='Synthèse', index=False)

        # After writing the data, apply formatting
        wb = load_workbook(excel_file)

        # Apply formatting to each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Define colors and fonts
            header_fill = PatternFill(start_color="009999", end_color="009999", fill_type="solid")
            header_font = Font(color="FFFFFF", size=10)
            cell_font = Font(color="000000", size=9)
            italic_font = Font(italic=True, color="000000")
            border_color = "D9D9D9"
            thin_border = Border(
                left=Side(style='thin', color=border_color),
                right=Side(style='thin', color=border_color),
                top=Side(style='thin', color=border_color),
                bottom=Side(style='thin', color=border_color)
            )

            # Apply header formatting
            for cell in ws[1]:
                # Set header font to Calibri, bold and color white
                cell.font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply cell formatting
            for row in ws.iter_rows(min_row=2):  # iter_rows() directly returns the rows, no need for len()
                for cell in row:
                    cell.font = cell_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = thin_border

                    # If the cell contains an 'X', highlight it with a green background
                    if cell.value == 'X':
                        cell.fill = PatternFill(start_color="91d2ff", end_color="91d2ff", fill_type="solid")

            # Hide all sheets except 'Synthèse'
            if sheet_name != 'Synthèse':
                ws.sheet_state = 'hidden'

        # Save the formatted workbook
        wb.save(excel_file)

        print(f"Data successfully exported and formatted to {excel_file}")

    except Exception as e:
        print(f"Error writing to Excel file: {e}")

# Demander à l'utilisateur de spécifier le chemin du dossier contenant les fichiers XML
folder_path = obtain_folder_path()

# Appeler la fonction avec le chemin du dossier fourni par l'utilisateur
process_xml_files_in_folder(folder_path)
