"""
Auteur : ExEco Environnement - François Botcazou
Date de création : 2025/02
Dernière mise à jour : 2025/03
Version : 1.0
Nom : ZnieffXmlToXlsxHab.py
Groupe : Biblizou_PatNat
Description : Module pour extraire des données d'habitats déterminants à partir de fichiers XML et les exporter dans un fichier Excel.
Dépendances :
    - Python 3.x
    - QGIS (QgsMessageLog)
    - xml.etree.ElementTree
    - pandas
    - openpyxl
    - os, datetime

Utilisation :
    Ce module doit être appelé depuis une extension QGIS.
"""

from qgis.core import QgsMessageLog
from qgis.gui import QgsMessageBar
from PyQt5.QtWidgets import QFileDialog, QDialog
import xml.etree.ElementTree as ET
import pandas as pd
import os
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl import load_workbook
from datetime import datetime


class ZnieffXmlToXlsxHab:
    def __init__(self, iface):
        self.iface = iface

    def run(self):
        # Demander à l'utilisateur de choisir un dossier contenant les fichiers XML
        folder_path = QFileDialog.getExistingDirectory(None, "Sélectionner un dossier contenant les fichiers XML")
        if not folder_path:
            self.iface.messageBar().pushMessage("Annulation", "Aucun dossier sélectionné.", level=Qgis.Info)
            return

        self.process_xml_files_in_folder(folder_path)

    def truncate_sheet_name(self, sheet_name):
        return sheet_name[:31]

    def xml_to_dataframe(self, xml_file):
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            lb_codes, lb_habs, lb_zn, nm_sffzn = [], [], "", ""

            for znieff_elem in root.iter('ZNIEFF'):
                nm_sffzn_elem = znieff_elem.find('NM_SFFZN')
                lb_zn_elem = znieff_elem.find('LB_ZN')
                if nm_sffzn_elem is not None:
                    nm_sffzn = nm_sffzn_elem.text
                if lb_zn_elem is not None:
                    lb_zn = lb_zn_elem.text
                for typo_info_row_elem in znieff_elem.iter('TYPO_INFO_ROW'):
                    fg_typo_elem = typo_info_row_elem.find('FG_TYPO')
                    if fg_typo_elem is not None and fg_typo_elem.text == 'D':
                        for lb_code_elem in typo_info_row_elem.iter('LB_CODE'):
                            lb_codes.append(lb_code_elem.text)
                        for lb_hab_elem in typo_info_row_elem.iter('LB_HAB'):
                            lb_habs.append(lb_hab_elem.text)

            data = {'LB_CODE': lb_codes, 'LB_HAB': lb_habs}
            df = pd.DataFrame(data)
            return df, lb_zn, nm_sffzn
        except ET.ParseError as e:
            QgsMessageLog.logMessage(f"Erreur de parsing XML: {xml_file} - {e}", "Biblizou", level=Qgis.Critical)
            return pd.DataFrame(columns=['LB_CODE', 'LB_HAB']), "", ""
        except Exception as e:
            QgsMessageLog.logMessage(f"Erreur inattendue avec {xml_file}: {e}", "Biblizou", level=Qgis.Critical)
            return pd.DataFrame(columns=['LB_CODE', 'LB_HAB']), "", ""

    def process_xml_files_in_folder(self, folder_path):
        if not os.path.isdir(folder_path):
            self.iface.messageBar().pushMessage("Erreur", "Le chemin sélectionné n'est pas un dossier valide.",
                                                level=Qgis.Warning)
            return

        xml_files = [f for f in os.listdir(folder_path) if
                     f.endswith('.xml') and not f.startswith('FR') and len(f) == 13]
        unique_lb_codes, unique_lb_habs = set(), set()
        hab_presence = {}
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        excel_file = os.path.join(folder_path, f'ZNIEFF_synthèse_des_habitats_déterminants_{current_time}.xlsx')

        try:
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                for xml_file in xml_files:
                    full_path = os.path.join(folder_path, xml_file)
                    df, lb_zn, nm_sffzn = self.xml_to_dataframe(full_path)
                    if not df.empty:
                        unique_lb_codes.update(df['LB_CODE'].unique())
                        unique_lb_habs.update(df['LB_HAB'].unique())
                        sheet_name = f"{nm_sffzn} - {lb_zn}"
                        sheet_name_truncated = self.truncate_sheet_name(sheet_name)
                        hab_presence[sheet_name_truncated] = set(df['LB_HAB'])
                        df.to_excel(writer, sheet_name=sheet_name_truncated, index=False)
                summary_data = {'LB_CODE': list(unique_lb_codes), 'LB_HAB': list(unique_lb_habs)}
                for sheet_name in hab_presence:
                    summary_data[sheet_name] = ['X' if hab in hab_presence[sheet_name] else '' for hab in
                                                summary_data['LB_HAB']]
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Synthèse', index=False)
            wb = load_workbook(excel_file)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for cell in ws[1]:
                    cell.font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
                    cell.fill = PatternFill(start_color="009999", end_color="009999", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = Font(color="000000", size=9)
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.border = Border(left=Side(style='thin', color="D9D9D9"),
                                             right=Side(style='thin', color="D9D9D9"),
                                             top=Side(style='thin', color="D9D9D9"),
                                             bottom=Side(style='thin', color="D9D9D9"))
                        if cell.value == 'X':
                            cell.fill = PatternFill(start_color="91d2ff", end_color="91d2ff", fill_type="solid")
                if sheet_name != 'Synthèse':
                    ws.sheet_state = 'hidden'
            wb.save(excel_file)
            self.iface.messageBar().pushMessage("Succès", f"Données exportées dans {excel_file}", level=Qgis.Success)
        except Exception as e:
            self.iface.messageBar().pushMessage("Erreur", f"Problème lors de l'export Excel : {e}", level=Qgis.Critical)
