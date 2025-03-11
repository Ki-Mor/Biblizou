"""
Author : ExEco Environnement
Edition date : 2025/02
Name : 13_znieff_xml2xlsx_esp
Group : Biblio_PatNat
"""

import xml.etree.ElementTree as ET
from xml.etree.ElementTree import Element
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from datetime import datetime
import sys
import time

def obtain_folder_path():
    return sys.argv[1] if len(sys.argv) > 1 else input("Veuillez entrer le chemin du dossier : ")

def truncate_sheet_name(sheet_name):
    """
    Truncate the sheet name to 31 characters, as required by Excel.
    """
    return sheet_name[:31]  # Truncate to 31 characters

def xml_to_dataframe(xml_file):
    """
    Parse the XML file and return a DataFrame with the contents of the <REGNE>, <GROUPE>, <CD_NOM>,
    <NOM_COMPLET>, <NOM_VERN> tags inside the <ZNIEFF> elements, but only for rows with <FG_ESP> == 'D'
    followed by <REGNE>, <GROUPE>, <CD_NOM>, <NOM_COMPLET>, and <NOM_VERN>.
    """
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()

        # Initialize lists to store the data
        regnes,  groupes, cd_noms, nom_complets, nom_vern, lb_zn, nm_sffzn= [], [], [], [], [], "",""

        start_time: float = time.time()

        for znieff_elem in root.iter('ZNIEFF'):
            nm_sffzn_elem = znieff_elem.find('NM_SFFZN')
            lb_zn_elem = znieff_elem.find('LB_ZN')

            if nm_sffzn_elem is not None:
                nm_sffzn = nm_sffzn_elem.text  # Get NM_SFFZN value

            if lb_zn_elem is not None:
                lb_zn = lb_zn_elem.text  # Get LB_ZN value

            # Iterate through the <ESPECE_ROW> elements
            espece_row_elem: Element
            for espece_row_elem in znieff_elem.iter('ESPECE_ROW'):
                fg_esp_elem = espece_row_elem.find('FG_ESP')

                # Only consider the data if <FG_ESP> has the value 'D'
                if fg_esp_elem is not None and fg_esp_elem.text == 'D':
                    # Extract REGNE, GROUPE, CD_NOM, NOM_COMPLET, and NOM_VERN for the current <ESPECE_ROW>
                    regne_elem = espece_row_elem.find('REGNE')
                    groupe_elem = espece_row_elem.find('GROUPE')
                    cd_nom_elem = espece_row_elem.find('CD_NOM')
                    nom_complet_elem = espece_row_elem.find('NOM_COMPLET')
                    nom_vern_elem = espece_row_elem.find('NOM_VERN')

                    # Add data to the lists or append empty string if data is missing
                    regnes.append(regne_elem.text if regne_elem is not None else "")
                    groupes.append(groupe_elem.text if groupe_elem is not None else "")
                    cd_noms.append(cd_nom_elem.text if cd_nom_elem is not None else "")
                    nom_complets.append(nom_complet_elem.text if nom_complet_elem is not None else "")
                    nom_vern.append(nom_vern_elem.text if nom_vern_elem is not None else "")

        xml_processing_time = time.time() - start_time
        print(f"Fichier {xml_file} traité en {xml_processing_time:.2f} secondes.")

        # Create a DataFrame from the extracted data
        data = {
            'REGNE': regnes,
            'GROUPE': groupes,
            'CD_NOM': cd_noms,
            'NOM_COMPLET': nom_complets,
            'NOM_VERN': nom_vern
        }

        # Ensure that all lists have the same length by filling in missing values with ""
        max_len = max(len(regnes), len(groupes), len(cd_noms), len(nom_complets), len(nom_vern))
        regnes += [""] * (max_len - len(regnes))
        groupes += [""] * (max_len - len(groupes))
        cd_noms += [""] * (max_len - len(cd_noms))
        nom_complets += [""] * (max_len - len(nom_complets))
        nom_vern += [""] * (max_len - len(nom_vern))

        df = pd.DataFrame(data)
        return df, lb_zn, nm_sffzn  # Return LB_ZN and NM_SFFZN along with the DataFrame

    except ET.ParseError as e:
        print(f"Error parsing XML file {xml_file}: {e}")
        return pd.DataFrame(columns=['REGNE', 'GROUPE', 'CD_NOM', 'NOM_COMPLET', 'NOM_VERN']), "", ""
    except Exception as e:
        print(f"Unexpected error with file {xml_file}: {e}")
        return pd.DataFrame(columns=['REGNE', 'GROUPE', 'CD_NOM', 'NOM_COMPLET', 'NOM_VERN']), "", ""

def process_xml_files_in_folder(folder_path):
    """
    Process all XML files in a folder and export the data to an Excel file with a summary sheet for each NM_SFFZN.
    """
    # Validate the folder path
    if not os.path.isdir(folder_path):
        print(f"The path {folder_path} is not a valid directory.")
        return

    # List all XML files in the folder
    xml_files = [f for f in os.listdir(folder_path) if f.endswith('.xml') and not f.startswith('FR') and len(f) == 13]

    # Initialize dictionaries to store data for each NM_SFFZN
    nm_sffzn_data = {}

    # Initialize lists to store data for the "Animalia" and "Plantae" summary sheets
    animalia_data = []
    plantae_data = []

    # Get the current date and time for the filename
    current_time = datetime.now().strftime("%Y%m%d%H%M%S")

    # Define the output Excel file path with the timestamp
    excel_file = os.path.join(folder_path, f'ZNIEFF_synthèse_des_esp_déterminantes_{current_time}.xlsx')

    # Create an Excel writer object
    try:
        print(f"\nDébut du traitement des fichiers XML...")

        start_time = time.time()

        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            for idx, xml_file in enumerate(xml_files, start=1):
                print(f"Traitement du fichier {idx}/{len(xml_files)} : {xml_file}...")
                full_path = os.path.join(folder_path, xml_file)
                df, nm_sffzn, lb_zn = xml_to_dataframe(full_path)

                if not df.empty:
                    # Store the data for each NM_SFFZN
                    if nm_sffzn not in nm_sffzn_data:
                        nm_sffzn_data[nm_sffzn] = []
                    nm_sffzn_data[nm_sffzn].append(df)

                    # Add to Animalia data if REGNE == "Animalia"
                    if "Animalia" in df['REGNE'].values:
                        for _, row in df[df['REGNE'] == "Animalia"].iterrows():
                            animalia_data.append({
                                "GROUPE": row["GROUPE"],
                                "CD_NOM": row["CD_NOM"],
                                "NOM_COMPLET": row["NOM_COMPLET"],
                                "NOM_VERN": row["NOM_VERN"],
                                "NM_SFFZN - LB_ZN": f"{nm_sffzn} - {lb_zn}",
                            })

                    # Add to Plantae data if REGNE == "Plantae"
                    if "Plantae" in df['REGNE'].values:
                        for _, row in df[df['REGNE'] == "Plantae"].iterrows():
                            plantae_data.append({
                                "GROUPE": row["GROUPE"],
                                "CD_NOM": row["CD_NOM"],
                                "NOM_COMPLET": row["NOM_COMPLET"],
                                "NOM_VERN": row["NOM_VERN"],
                                "NM_SFFZN - LB_ZN": f"{nm_sffzn} - {lb_zn}",
                            })

            # Create the individual sheets for each NM_SFFZN
            for nm_sffzn, dfs in nm_sffzn_data.items():
                combined_df = pd.concat(dfs, ignore_index=True)

                # Sort the combined data by GROUPE, then by NOM_COMPLET (in reverse order)
                combined_df = combined_df.sort_values(by=['GROUPE', 'NOM_COMPLET'], ascending=[True, False])

                # Concatenate NM_SFFZN and LB_ZN for the sheet name
                print({nm_sffzn}, {lb_zn})
                sheet_name = f"{nm_sffzn} - {lb_zn}"
                print(sheet_name)
                sheet_name_truncated = truncate_sheet_name(sheet_name)

                # Create the sheet for this NM_SFFZN
                combined_df.to_excel(writer, sheet_name=sheet_name_truncated, index=False)

            # Create the "Synthèse Animalia" sheet if there is data
            if animalia_data:
                animalia_df = pd.DataFrame(animalia_data)

                pivot_df = animalia_df.pivot_table(index=["GROUPE", "CD_NOM", "NOM_COMPLET", "NOM_VERN"],
                                                   columns="NM_SFFZN - LB_ZN",
                                                   aggfunc=lambda x: "X", fill_value="")

                pivot_df.to_excel(writer, sheet_name="Synthèse Animalia", index=True)

            # Create the "Synthèse Plantae" sheet if there is data
            if plantae_data:
                plantae_df = pd.DataFrame(plantae_data)

                pivot_df = plantae_df.pivot_table(index=["GROUPE", "CD_NOM", "NOM_COMPLET", "NOM_VERN"],
                                                   columns="NM_SFFZN - LB_ZN",
                                                   aggfunc=lambda x: "X", fill_value="")

                pivot_df.to_excel(writer, sheet_name="Synthèse Plantae", index=True)

        processing_time = time.time() - start_time
        print(f"\nTraitement des fichiers XML terminé en {processing_time:.2f} secondes.")


        wb = load_workbook(excel_file)

        # Apply formatting to each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            header_fill = PatternFill(start_color="009999", end_color="009999", fill_type="solid")
            header_font = Font(color="FFFFFF", size=10)
            cell_font = Font(color="000000", size=9)
            border_color = "D9D9D9"
            thin_border = Border(
                left=Side(style='thin', color=border_color),
                right=Side(style='thin', color=border_color),
                top=Side(style='thin', color=border_color),
                bottom=Side(style='thin', color=border_color)
            )

            for cell in ws[1]:
                # Set header font to Calibri, bold and color white
                cell.font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in ws.iter_rows(min_row=2):  # iter_rows() directly returns the rows, no need for len()
                for cell in row:
                    cell.font = cell_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = thin_border

                    if cell.value == 'X':
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color="91d2ff", end_color="91d2ff", fill_type="solid")

            if sheet_name not in ["Synthèse Animalia", "Synthèse Plantae"]:
                ws.sheet_state = 'hidden'

        wb.save(excel_file)

        print(f"Data successfully exported and formatted to {excel_file}")

    except Exception as e:
        print(f"Error writing to Excel file: {e}")

folder_path = obtain_folder_path()
process_xml_files_in_folder(folder_path)
