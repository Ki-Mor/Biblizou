"""
Auteur : ExEco Environnement - François Botcazou
Date de création : 2025/02
Dernière mise à jour : 2025/03
Version : 1.0
Nom : ZnieffXmlToXlsxHab.py
Groupe : Biblizou_PatNat
Description : Module pour extraire des données d'habitats déterminants à partir de fichiers XML et les exporter dans un fichier Excel.
Dépendances :
    - Python 3.x
    - QGIS (QgsMessageLog)
    - xml.etree.ElementTree
    - pandas
    - openpyxl
    - os, datetime

Utilisation :
    Ce module doit être appelé depuis une extension QGIS.
"""

import xml.etree.ElementTree as ET
import pandas as pd
import os
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl import load_workbook
from datetime import datetime
from qgis.core import QgsMessageLog


class ZnieffXmlToXlsxHab:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.excel_file = self._generate_output_filepath()

    def _generate_output_filepath(self):
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        return os.path.join(self.folder_path, f'ZNIEFF_synthèse_des_habitats_déterminants_{current_time}.xlsx')

    def _truncate_sheet_name(self, sheet_name):
        return sheet_name[:31]

    def _xml_to_dataframe(self, xml_file):
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            lb_codes, lb_habs, lb_zn, nm_sffzn = [], [], "", ""

            for znieff_elem in root.iter('ZNIEFF'):
                nm_sffzn_elem = znieff_elem.find('NM_SFFZN')
                lb_zn_elem = znieff_elem.find('LB_ZN')
                if nm_sffzn_elem is not None:
                    nm_sffzn = nm_sffzn_elem.text
                if lb_zn_elem is not None:
                    lb_zn = lb_zn_elem.text

                for typo_info_row_elem in znieff_elem.iter('TYPO_INFO_ROW'):
                    fg_typo_elem = typo_info_row_elem.find('FG_TYPO')
                    if fg_typo_elem is not None and fg_typo_elem.text == 'D':
                        for lb_code_elem in typo_info_row_elem.iter('LB_CODE'):
                            lb_codes.append(lb_code_elem.text)
                        for lb_hab_elem in typo_info_row_elem.iter('LB_HAB'):
                            lb_habs.append(lb_hab_elem.text)

            df = pd.DataFrame({'LB_CODE': lb_codes, 'LB_HAB': lb_habs})
            return df, lb_zn, nm_sffzn
        except ET.ParseError as e:
            QgsMessageLog.logMessage(f"Erreur de parsing XML: {xml_file} - {e}", "ZnieffXmlToXlsxHab", 2)
        except Exception as e:
            QgsMessageLog.logMessage(f"Erreur inattendue avec {xml_file}: {e}", "ZnieffXmlToXlsxHab", 2)
        return pd.DataFrame(columns=['LB_CODE', 'LB_HAB']), "", ""

    def process_files(self):
        if not os.path.isdir(self.folder_path):
            QgsMessageLog.logMessage(f"Chemin invalide: {self.folder_path}", "ZnieffXmlToXlsxHab", 2)
            return

        xml_files = [f for f in os.listdir(self.folder_path) if
                     f.endswith('.xml') and not f.startswith('FR') and len(f) == 13]
        unique_lb_codes, unique_lb_habs, hab_presence = set(), set(), {}

        with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
            for xml_file in xml_files:
                df, lb_zn, nm_sffzn = self._xml_to_dataframe(os.path.join(self.folder_path, xml_file))
                if not df.empty:
                    unique_lb_codes.update(df['LB_CODE'].unique())
                    unique_lb_habs.update(df['LB_HAB'].unique())
                    sheet_name = self._truncate_sheet_name(f"{nm_sffzn} - {lb_zn}")
                    hab_presence[sheet_name] = set(df['LB_HAB'])
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            summary_data = {'LB_CODE': list(unique_lb_codes), 'LB_HAB': list(unique_lb_habs)}
            for sheet_name in hab_presence:
                summary_data[sheet_name] = ['X' if hab in hab_presence[sheet_name] else '' for hab in
                                            summary_data['LB_HAB']]

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Synthèse', index=False)

        self._format_excel()
        QgsMessageLog.logMessage(f"Export terminé: {self.excel_file}", "ZnieffXmlToXlsxHab", 0)

    def _format_excel(self):
        wb = load_workbook(self.excel_file)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_fill = PatternFill(start_color="009999", fill_type="solid")
            cell_font = Font(size=9)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = cell_font
                    cell.alignment = Alignment(horizontal='left')
                    cell.border = thin_border
                    if cell.value == 'X':
                        cell.fill = PatternFill(start_color="91d2ff", fill_type="solid")

            if sheet_name != 'Synthèse':
                ws.sheet_state = 'hidden'

        wb.save(self.excel_file)

# Exemple d'utilisation dans QGIS
# processor = ZnieffXmlToXlsxHab("/chemin/vers/dossier")
# processor.process_files()
