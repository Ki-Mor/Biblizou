"""
Author : ExEco Environnement
Edition date : 2025/02
Name : 23_n2000_xml2xlsx_esp
Group : Biblio_PatNat
"""

import xml.etree.ElementTree as ET
import requests
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
import sys
import time
from collections import defaultdict

def obtain_folder_path():
    return sys.argv[1] if len(sys.argv) > 1 else input("Veuillez entrer le chemin du dossier : ")

def truncate_sheet_name(sheet_name):
    return sheet_name[:31]  # Limiter à 31 caractères maximum (restriction d'Excel)

def get_taxref_data(cd_nom, cache={}):
    # Vérifier si les données sont déjà en cache
    if cd_nom in cache:
        return cache[cd_nom]

    # URL de l'API avec cd_nom comme identifiant
    url = f"https://taxref.mnhn.fr/api/taxa/{cd_nom}"

    # Définir les headers pour l'API
    headers = {"accept": "application/hal+json;version=1"}

    try:
        # Faire la requête GET
        response = requests.get(url, headers=headers)

        # Vérifier que la réponse est correcte (code 200)
        if response.status_code == 200:
            data = response.json()

            # Vérifier si la réponse contient des informations
            if data:
                # Récupérer les informations pertinentes directement dans la réponse
                result = {
                    'REGNE': data.get('kingdomName', ''),
                    'GROUPE': data.get('vernacularGroup2', ''),
                    'NOM_COMPLET': data.get('fullName', ''),
                    'NOM_VERN': data.get('frenchVernacularName', '')
                }

                # Mettre les données dans le cache
                cache[cd_nom] = result
                return result
            else:
                print(f"Aucun taxon trouvé pour {cd_nom}.")
                return {'REGNE': '', 'GROUPE': '', 'NOM_COMPLET': '', 'NOM_VERN': ''}
        else:
            print(f"Erreur API pour {cd_nom}: {response.status_code}")
            return {'REGNE': '', 'GROUPE': '', 'NOM_COMPLET': '', 'NOM_VERN': ''}
    except Exception as e:
        print(f"Erreur lors de l'interrogation de l'API pour {cd_nom}: {e}")
        return {'REGNE': '', 'GROUPE': '', 'NOM_COMPLET': '', 'NOM_VERN': ''}

def xml_to_dataframe(xml_file, cache):
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()

        regnes, groupes, cd_noms, noms, nom_complets, nom_vern, sitecode, site_name = [], [], [], [], [], [], "", ""

        start_time: float = time.time()

        # Recherche de l'élément BIOTOP
        for biotop_elem in root.iter('BIOTOP'):
            sitecode_elem = biotop_elem.find('SITECODE')
            site_name_elem = biotop_elem.find('SITE_NAME')

            # Vérification si les éléments sont bien trouvés et non vides
            if sitecode_elem is not None and sitecode_elem.text:
                sitecode = sitecode_elem.text.strip()
            else:
                sitecode = None  # ou 'no_sitecode' si tu veux une valeur par défaut

            if site_name_elem is not None and site_name_elem.text:
                site_name = site_name_elem.text.strip()
            else:
                site_name = None  # ou 'no_site_name' si tu veux une valeur par défaut

            for species_elem in biotop_elem.iter('SPECIES'):
                for species_row_elem in species_elem.iter('SPECIES_ROW'):
                    nom_elem = species_row_elem.find('NOM')
                    cd_nom_elem = species_row_elem.find('CD_NOM')

                    if cd_nom_elem is not None:
                        cd_nom = cd_nom_elem.text
                        nom = nom_elem.text
                        taxon_info = get_taxref_data(cd_nom, cache)

                        regnes.append(taxon_info['REGNE'])
                        groupes.append(taxon_info['GROUPE'])
                        cd_noms.append(cd_nom_elem.text if cd_nom_elem is not None else "")
                        noms.append(nom)
                        nom_complets.append(taxon_info['NOM_COMPLET'])
                        nom_vern.append(taxon_info['NOM_VERN'])

        xml_processing_time = time.time() - start_time
        print(f"Fichier {xml_file} traité en {xml_processing_time:.2f} secondes.")

        # Retourner le DataFrame avec les bonnes données
        data = {
            'REGNE': regnes,
            'GROUPE': groupes,
            'CD_NOM': cd_noms,
            'NOM': noms,
            'NOM_COMPLET': nom_complets,
            'NOM_VERN': nom_vern
        }

        # S'assurer que toutes les listes ont la même longueur
        max_len = max(len(regnes), len(groupes), len(cd_noms), len(noms), len(nom_complets), len(nom_vern))
        regnes += [""] * (max_len - len(regnes))
        groupes += [""] * (max_len - len(groupes))
        cd_noms += [""] * (max_len - len(cd_noms))
        noms += [""] * (max_len - len(noms))
        nom_complets += [""] * (max_len - len(nom_complets))
        nom_vern += [""] * (max_len - len(nom_vern))

        df = pd.DataFrame(data)
        return df, sitecode, site_name

    except ET.ParseError as e:
        print(f"Error parsing XML file {xml_file}: {e}")
        return pd.DataFrame(columns=['REGNE', 'GROUPE', 'CD_NOM', 'NOM', 'NOM_COMPLET', 'NOM_VERN']), "", ""
    except Exception as e:
        print(f"Unexpected error with file {xml_file}: {e}")
        return pd.DataFrame(columns=['REGNE', 'GROUPE', 'CD_NOM', 'NOM', 'NOM_COMPLET', 'NOM_VERN']), "", ""

def process_xml_files_in_folder(folder_path):
    if not os.path.isdir(folder_path):
        print(f"The path {folder_path} is not a valid directory.")
        return

    xml_files = [f for f in os.listdir(folder_path) if f.startswith('FR') and f.endswith('.xml') and len(f) == 13]

    if not xml_files:
        print(f"No XML files found in the folder {folder_path}")
        return

    cache = {}

    # Initialize lists to store data for the "Animalia" and "Plantae" summary sheets
    animalia_data = []
    plantae_data = []

    current_time = datetime.now().strftime("%Y%m%d%H%M%S")
    excel_file = os.path.join(folder_path, f'output_n2000_xml2xlsx_esp_{current_time}.xlsx')

    try:
        print(f"\nDébut du traitement des fichiers XML...")

        start_time = time.time()

        # Utilisation de pd.ExcelWriter pour écrire dans le fichier Excel
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            for idx, xml_file in enumerate(xml_files, start=1):
                print(f"Traitement du fichier {idx}/{len(xml_files)} : {xml_file}...")
                full_path = os.path.join(folder_path, xml_file)
                df, sitecode, site_name = xml_to_dataframe(full_path, cache)

                if not df.empty:
                    # Vérifier et remplacer les valeurs vides ou manquantes de SITECODE et SITE_NAME
                    if not sitecode:
                        print(f"Warning: SITECODE is empty for file {xml_file}. Using 'no_sitecode'.")
                        sitecode = "no_sitecode"
                    if not site_name:
                        print(f"Warning: SITE_NAME is empty for file {xml_file}. Using 'no_site_name'.")
                        site_name = "no_site_name"

                    # Ajouter à la liste animalia_data si REGNE == "Animalia"
                    if "Animalia" in df['REGNE'].values:
                        for _, row in df[df['REGNE'] == "Animalia"].iterrows():
                            animalia_data.append({
                                "GROUPE": row["GROUPE"],
                                "CD_NOM": row["CD_NOM"],
                                "NOM_COMPLET": row["NOM_COMPLET"],
                                "NOM_VERN": row["NOM_VERN"],
                                "SITECODE - SITE_NAME": f"{sitecode} - {site_name}",
                            })

                    # Ajouter à la liste plantae_data si REGNE == "Plantae"
                    if "Plantae" in df['REGNE'].values:
                        for _, row in df[df['REGNE'] == "Plantae"].iterrows():
                            plantae_data.append({
                                "GROUPE": row["GROUPE"],
                                "CD_NOM": row["CD_NOM"],
                                "NOM_COMPLET": row["NOM_COMPLET"],
                                "NOM_VERN": row["NOM_VERN"],
                                "SITECODE - SITE_NAME": f"{sitecode} - {site_name}",
                            })

                    # Générer un nom d'onglet unique basé sur SITECODE et SITE_NAME
                    sheet_name = f"{sitecode}-{site_name}"
                    print(f"Nom d'onglet généré: {sheet_name}")

                    sheet_name_truncated = truncate_sheet_name(sheet_name)

                    # Écrire les données dans la feuille correspondante
                    df.drop(columns=['NOM'], inplace=True)  # Exclure la colonne NOM
                    df.to_excel(writer, sheet_name=sheet_name_truncated, index=False)
                    print(f"Le fichier {xml_file} a été traité et ajouté sous la feuille {sheet_name_truncated}.")
                else:
                    print(f"Warning: No data exported for {xml_file}")

        # Créer la feuille "Synthèse Animalia" après la boucle
        if animalia_data:
            animalia_df = pd.DataFrame(animalia_data)

            # Pivot the table
            pivot_df = animalia_df.pivot_table(index=["GROUPE", "CD_NOM", "NOM_COMPLET", "NOM_VERN"],
                                               columns="SITECODE - SITE_NAME",
                                               aggfunc=lambda x: "X", fill_value="")
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
                pivot_df.to_excel(writer, sheet_name="Synthèse Animalia", index=True)
                print("La feuille 'Synthèse Animalia' a été créée.")

        # Créer la feuille "Synthèse Plantae" après la boucle
        if plantae_data:
            plantae_df = pd.DataFrame(plantae_data)

            # Pivot the table to get "X" where CD_NOM is associated with NM_SFFZN - LB_ZN
            pivot_df = plantae_df.pivot_table(index=["GROUPE", "CD_NOM", "NOM_COMPLET", "NOM_VERN"],
                                               columns="SITECODE - SITE_NAME",
                                               aggfunc=lambda x: "X", fill_value="")
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
                pivot_df.to_excel(writer, sheet_name="Synthèse Plantae", index=True)
                print("La feuille 'Synthèse Plantae' a été créée.")

        processing_time = time.time() - start_time
        print(f"\nTraitement des fichiers XML terminé en {processing_time:.2f} secondes.")

        # After writing the data, apply formatting
        wb = load_workbook(excel_file)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            header_fill = PatternFill(start_color="009999", end_color="009999", fill_type="solid")
            cell_font = Font(color="000000", size=9)
            border_color = "D9D9D9"
            thin_border = Border(
                left=Side(style='thin', color=border_color),
                right=Side(style='thin', color=border_color),
                top=Side(style='thin', color=border_color),
                bottom=Side(style='thin', color=border_color)
            )

            for cell in ws[1]:
                cell.font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in ws.iter_rows(min_row=2):  # iter_rows() directly returns the rows, no need for len()
                for cell in row:
                    cell.font = cell_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = thin_border

                    if cell.value == 'X':
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color="91d2ff", end_color="91d2ff", fill_type="solid")

            if sheet_name not in ["Synthèse Animalia", "Synthèse Plantae"]:
                ws.sheet_state = 'hidden'

        wb.save(excel_file)

        print(f"Data successfully exported and formatted to {excel_file}")

    except Exception as e:
        print(f"An error occurred while processing XML files: {e}")

folder_path = obtain_folder_path()
process_xml_files_in_folder(folder_path)
