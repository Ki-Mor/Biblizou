"""
Auteur : ExEco Environnement - François Botcazou
Date de création : 2025/02
Dernière mise à jour : 2025/03
Version : 1.0
Nom : ZnieffXmlToXlsxEsp.py
Groupe : Biblizou_PatNat
Description : Module pour extraire des données d'espèces déterminantes à partir de fichiers XML et les exporter dans un fichier Excel.
Dépendances :
    - Python 3.x
    - QGIS (QgsMessageLog)
    - xml.etree.ElementTree
    - pandas
    - openpyxl
    - os, datetime

Utilisation :
    Ce module doit être appelé depuis une extension QGIS.
"""
import xml.etree.ElementTree as ET
import pandas as pd
import os
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from datetime import datetime
from qgis.core import QgsMessageLog


class ZnieffXmlProcessor:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.excel_file = os.path.join(folder_path, f'ZNIEFF_synthese_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx')
        self.logger = QgsMessageLog

    def truncate_sheet_name(self, sheet_name):
        return sheet_name[:31]

    def xml_to_dataframe(self, xml_file):
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            data = {"REGNE": [], "GROUPE": [], "CD_NOM": [], "NOM_COMPLET": [], "NOM_VERN": []}
            lb_zn, nm_sffzn = "", ""

            for znieff_elem in root.iter('ZNIEFF'):
                lb_zn_elem, nm_sffzn_elem = znieff_elem.find('LB_ZN'), znieff_elem.find('NM_SFFZN')
                lb_zn, nm_sffzn = (lb_zn_elem.text if lb_zn_elem else ""), (nm_sffzn_elem.text if nm_sffzn_elem else "")

                for espece_row_elem in znieff_elem.iter('ESPECE_ROW'):
                    if espece_row_elem.find('FG_ESP') is not None and espece_row_elem.find('FG_ESP').text == 'D':
                        for key in data.keys():
                            elem = espece_row_elem.find(key)
                            data[key].append(elem.text if elem is not None else "")

            return pd.DataFrame(data), lb_zn, nm_sffzn
        except ET.ParseError as e:
            self.logger.logMessage(f"Erreur de parsing XML {xml_file}: {e}", level=2)
        except Exception as e:
            self.logger.logMessage(f"Erreur inattendue avec {xml_file}: {e}", level=2)
        return pd.DataFrame(columns=["REGNE", "GROUPE", "CD_NOM", "NOM_COMPLET", "NOM_VERN"]), "", ""

    def process_xml_files(self):
        if not os.path.isdir(self.folder_path):
            self.logger.logMessage(f"Dossier invalide: {self.folder_path}", level=2)
            return

        xml_files = [f for f in os.listdir(self.folder_path) if
                     f.endswith('.xml') and not f.startswith('FR') and len(f) == 13]
        nm_sffzn_data, animalia_data, plantae_data = {}, [], []

        self.logger.logMessage("Début du traitement des fichiers XML...")
        with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
            for xml_file in xml_files:
                df, lb_zn, nm_sffzn = self.xml_to_dataframe(os.path.join(self.folder_path, xml_file))
                if df.empty:
                    continue
                if nm_sffzn not in nm_sffzn_data:
                    nm_sffzn_data[nm_sffzn] = []
                nm_sffzn_data[nm_sffzn].append(df)

                for regne, collection in [("Animalia", animalia_data), ("Plantae", plantae_data)]:
                    if regne in df['REGNE'].values:
                        collection.extend(
                            df[df['REGNE'] == regne].assign(NM_SFFZN_LB_ZN=f"{nm_sffzn} - {lb_zn}").to_dict(
                                orient='records'))

            for nm_sffzn, dfs in nm_sffzn_data.items():
                combined_df = pd.concat(dfs, ignore_index=True).sort_values(by=['GROUPE', 'NOM_COMPLET'],
                                                                            ascending=[True, False])
                combined_df.to_excel(writer, sheet_name=self.truncate_sheet_name(f"{nm_sffzn} - {lb_zn}"), index=False)

            for name, data in [("Synthèse Animalia", animalia_data), ("Synthèse Plantae", plantae_data)]:
                if data:
                    pd.DataFrame(data).pivot_table(index=["GROUPE", "CD_NOM", "NOM_COMPLET", "NOM_VERN"],
                                                   columns="NM_SFFZN_LB_ZN", aggfunc=lambda x: "X",
                                                   fill_value="").to_excel(writer, sheet_name=name, index=True)

        self.format_excel()
        self.logger.logMessage(f"Données exportées avec succès vers {self.excel_file}")

    def format_excel(self):
        try:
            wb = load_workbook(self.excel_file)
            header_fill = PatternFill(start_color="009999", end_color="009999", fill_type="solid")
            header_font = Font(color="FFFFFF", size=10)
            cell_font = Font(color="000000", size=9)
            border_color = "D9D9D9"
            thin_border = Border(left=Side(style='thin', color=border_color),
                                 right=Side(style='thin', color=border_color),
                                 top=Side(style='thin', color=border_color),
                                 bottom=Side(style='thin', color=border_color))

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for cell in ws[1]:
                    cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal='center',
                                                                                               vertical='center')
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font, cell.alignment, cell.border = cell_font, Alignment(horizontal='left',
                                                                                      vertical='center'), thin_border
                        if cell.value == 'X':
                            cell.alignment, cell.fill = Alignment(horizontal='center', vertical='center'), PatternFill(
                                start_color="91d2ff", end_color="91d2ff", fill_type="solid")
                if sheet_name not in ["Synthèse Animalia", "Synthèse Plantae"]:
                    ws.sheet_state = 'hidden'
            wb.save(self.excel_file)
        except Exception as e:
            self.logger.logMessage(f"Erreur de mise en forme Excel: {e}", level=2)
